import os
import sys
import io
import time
import shutil
import requests
import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import zipfile
import docx
from docx2pdf import convert

# Đảm bảo in tiếng Việt không lỗi trên terminal
if not hasattr(sys.stdout, 'original_stdout'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import json

# ==========================================
# 1. CẤU HÌNH API LARK MẶC ĐỊNH
# ==========================================
DEFAULT_APP_ID = "cli_aa8847bf0f79deef"
DEFAULT_APP_SECRET = "zBA9rmZ7ScgEtIS1OFFOedGIHaz11V8g"
DEFAULT_APP_TOKEN = "UWhibaHkCaWLausconRl8krsgZg"
DEFAULT_TABLE_CNTT = "tblZoFTIttbvejQ7"
DEFAULT_TABLE_QTKD = "tbl4gHCX7H8eGjaA"
DEFAULT_VIEW_CNTT = "vewHEt0hna"
DEFAULT_VIEW_QTKD = "vewuctEM09"

APP_ID = DEFAULT_APP_ID
APP_SECRET = DEFAULT_APP_SECRET
APP_TOKEN = DEFAULT_APP_TOKEN
TABLE_CNTT = DEFAULT_TABLE_CNTT
TABLE_QTKD = DEFAULT_TABLE_QTKD
VIEW_CNTT = DEFAULT_VIEW_CNTT
VIEW_QTKD = DEFAULT_VIEW_QTKD

# Đọc cấu hình từ config.json nếu tồn tại (để Web UI ghi đè)
try:
    if os.path.exists('config.json'):
        with open('config.json', 'r', encoding='utf-8') as f:
            _config = json.load(f)
            APP_ID = _config.get('APP_ID', DEFAULT_APP_ID)
            APP_SECRET = _config.get('APP_SECRET', DEFAULT_APP_SECRET)
            APP_TOKEN = _config.get('APP_TOKEN', DEFAULT_APP_TOKEN)
            TABLE_CNTT = _config.get('TABLE_CNTT', DEFAULT_TABLE_CNTT)
            TABLE_QTKD = _config.get('TABLE_QTKD', DEFAULT_TABLE_QTKD)
            VIEW_CNTT = _config.get('VIEW_CNTT', DEFAULT_VIEW_CNTT)
            VIEW_QTKD = _config.get('VIEW_QTKD', DEFAULT_VIEW_QTKD)
except Exception as e:
    print(f"Cảnh báo đọc file config.json: {e}. Sử dụng cấu hình mặc định.")

FILE_FIELD_NAME = 'File trúng tuyển vòng 1'
SKIP_EXISTING = True  # Đổi thành True để bỏ qua những người đã có file, chỉ update người mới

# ==========================================
# 2. BỘ TỪ ĐIỂN VÀ HÀM LÀM SẠCH (CLEAN DATA)
# ==========================================
AUX_WORDS = sorted([
    'số nhà', 'thành phố', 'thị trấn', 'thị xã', 'khu phố', 'dân phố',
    'tỉnh', 'xã', 'thôn', 'xóm', 'phường', 'quận', 'huyện',
    'khối', 'tổ', 'ngõ', 'ngách', 'đường', 'phố',
    'số', 'khu', 'ấp', 'làng', 'hẻm', 'lô',
], key=len, reverse=True)

PLUS_CODE_RE = re.compile(r'\b[A-Z0-9]{4,8}\+[A-Z0-9]{2,4}\b', re.IGNORECASE)

ABBREV_MAP = [
    (re.compile(r'\bTP\.\s*', re.IGNORECASE), 'Thành phố '),
    (re.compile(r'\bTX\.\s*', re.IGNORECASE), 'Thị xã '),
    (re.compile(r'\bTT\.\s*', re.IGNORECASE), 'Thị trấn '),
    (re.compile(r'\bP\.\s*',  re.IGNORECASE), 'Phường '),
    (re.compile(r'\bQ\.\s*',  re.IGNORECASE), 'Quận '),
    (re.compile(r'\bH\.\s*',  re.IGNORECASE), 'Huyện '),
    (re.compile(r'\bX\.\s*',  re.IGNORECASE), 'Xã '),
    (re.compile(r'\bHN\b',    re.IGNORECASE), 'Hà Nội'),
    (re.compile(r'\bHCM\b',   re.IGNORECASE), 'Hồ Chí Minh'),
    (re.compile(r'\bTPHCM\b', re.IGNORECASE), 'Hồ Chí Minh'),
    (re.compile(r'\bĐN\b',    re.IGNORECASE), 'Đà Nẵng'),
]

PROVINCES = [
    'Hà Nội', 'Hồ Chí Minh', 'Hải Phòng', 'Đà Nẵng', 'Cần Thơ',
    'An Giang', 'Bà Rịa Vũng Tàu', 'Bắc Giang', 'Bắc Kạn', 'Bạc Liêu',
    'Bắc Ninh', 'Bến Tre', 'Bình Định', 'Bình Dương', 'Bình Phước',
    'Bình Thuận', 'Cà Mau', 'Cao Bằng', 'Đắk Lắk', 'Đắk Nông',
    'Điện Biên', 'Đồng Nai', 'Đồng Tháp', 'Gia Lai', 'Hà Giang',
    'Hà Nam', 'Hà Tĩnh', 'Hải Dương', 'Hậu Giang', 'Hòa Bình',
    'Hưng Yên', 'Khánh Hòa', 'Kiên Giang', 'Kon Tum', 'Lai Châu',
    'Lâm Đồng', 'Lạng Sơn', 'Lào Cai', 'Long An', 'Nam Định',
    'Nghệ An', 'Ninh Bình', 'Ninh Thuận', 'Phú Thọ', 'Phú Yên',
    'Quảng Bình', 'Quảng Nam', 'Quảng Ngãi', 'Quảng Ninh', 'Quảng Trị',
    'Sóc Trăng', 'Sơn La', 'Tây Ninh', 'Thái Bình', 'Thái Nguyên',
    'Thanh Hóa', 'Thừa Thiên Huế', 'Tiền Giang', 'Trà Vinh', 'Tuyên Quang',
    'Vĩnh Long', 'Vĩnh Phúc', 'Yên Bái',
]
_PROV_LOWER = {p.lower(): p for p in PROVINCES}

def title_viet(word):
    return word[0].upper() + word[1:] if word else word

def normalize_abbrev(text):
    for pattern, replacement in ABBREV_MAP:
        text = pattern.sub(replacement, text)
    return text

def parse_segment(text):
    text = text.strip()
    text_lower = text.lower()
    
    prefixes = ["thành phố", "thị xã", "thị trấn", "quận", "huyện", "phường", "xã", "tỉnh"]
    for p in prefixes:
        if text_lower.startswith(p + " "):
            return p, text[len(p):].strip()
            
    return None, text

def is_duplicate_field(part, field_val):
    if not field_val or not field_val.strip():
        return False
        
    part_clean = part.strip().lower()
    field_clean = field_val.strip().lower()
    
    if part_clean == field_clean:
        return True
        
    part_norm = normalize_abbrev(part_clean)
    field_norm = normalize_abbrev(field_clean)
    if part_norm == field_norm:
        return True
        
    p_part, s_part = parse_segment(part_norm)
    p_field, s_field = parse_segment(field_norm)
    
    if s_part.lower() == s_field.lower():
        if p_part == p_field or p_part is None or p_field is None:
            return True
            
    return False

def clean_street_address(raw_street, raw_ward, raw_district, raw_city):
    if not raw_street:
        return ""
        
    parts = [p.strip() for p in raw_street.split(',') if p.strip()]
    cleaned_parts = []
    
    for part in parts:
        is_dup = False
        for field in [raw_ward, raw_district, raw_city]:
            if field and is_duplicate_field(part, field):
                is_dup = True
                break
        if not is_dup:
            cleaned_parts.append(part)
            
    return ", ".join(cleaned_parts)

def clean_ten(val):
    if not val: return ''
    s = re.sub(r'\s+', ' ', re.sub(r'\d+', '', str(val))).strip()
    return ' '.join(w[0].upper() + w[1:].lower() if w else w for w in s.split())

def clean_cccd(val):
    if val is None: return ''
    try: s = str(int(round(float(str(val)))))
    except Exception: s = re.sub(r'\D', '', str(val))
    s = re.sub(r'\D', '', s)
    return s.zfill(12)[:12]

def clean_sdt(val):
    if val is None: return ''
    try: s = str(int(round(float(str(val)))))
    except Exception: s = re.sub(r'\D', '', str(val))
    s = re.sub(r'\D', '', s)
    if len(s) == 9: s = '0' + s
    return s.zfill(10)[-10:]

def clean_ngaysinh_val(val):
    if not val: return ''
    try:
        if str(val).isdigit() and len(str(val)) > 10:
            dt = datetime.datetime.fromtimestamp(int(val)/1000)
            return dt.strftime('%d/%m/%Y')
        
        val_str = str(val).strip()
        date_part = val_str.split('T')[0].split(' ')[0]
        
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y'):
            try:
                dt = datetime.datetime.strptime(date_part, fmt)
                return dt.strftime('%d/%m/%Y')
            except Exception:
                continue
        return val_str
    except Exception:
        return str(val)

def normalize_vietnamese_tones(text):
    tone_map = {
        'oà': 'òa', 'oá': 'óa', 'oả': 'ỏa', 'oã': 'oã', 'oạ': 'ọa',
        'uý': 'úy', 'uỳ': 'ùy', 'uỷ': 'ủy', 'uỹ': 'uỹ', 'uỵ': 'ụy',
        'oè': 'òe', 'oé': 'óe', 'oẻ': 'ỏe', 'oẽ': 'õe', 'oẹ': 'ọe',
        'uế': 'uế',
    }
    for k, v in tone_map.items():
        text = text.replace(k, v)
    return text

def remove_vietnamese_accents(text):
    text = text.lower()
    accent_map = {
        'a': 'àáảãạăằắẳẵặâầấẩẫậ',
        'd': 'đ',
        'e': 'èéẻẽẹêềếểễệ',
        'i': 'ìíỉĩị',
        'o': 'òóỏõọôồốổỗộơờớởỡợ',
        'u': 'ùúủũụưừứửữự',
        'y': 'ỳýỷỹỵ',
    }
    for char, accented in accent_map.items():
        for acc in accented:
            text = text.replace(acc, char)
    return text

def make_accent_insensitive_regex(text):
    char_map = {
        'a': '[aàáảãạăằắẳẵặâầấẩẫậ]',
        'd': '[dđ]',
        'e': '[eèéẻẽẹêềếểễệ]',
        'i': '[iìíỉĩị]',
        'o': '[oòóỏõọôồốổỗộơờớởỡợ]',
        'u': '[uùúủũụưừứửữự]',
        'y': '[yỳýỷỹỵ]',
    }
    text = text.lower()
    text = remove_vietnamese_accents(text)
    res = []
    for char in text:
        if char in char_map:
            res.append(char_map[char])
        elif char.isspace():
            res.append(r'\s*')
        else:
            res.append(re.escape(char))
    return "".join(res)

def normalize_for_comparison(text):
    if not text:
        return ""
    text = text.lower()
    text = normalize_vietnamese_tones(text)
    text = normalize_abbrev(text)
    text = re.sub(r'\s+', ' ', text).strip()
    text = text.strip('.,;!-– ')
    return text

def remove_target_from_text(text, target):
    t_norm = normalize_for_comparison(target)
    if not t_norm or len(t_norm) < 3:
        return text
        
    _, core = parse_segment(t_norm)
    core = core.strip()
    
    core_regex = make_accent_insensitive_regex(core)
    prefix_regex = r'(?:tỉnh|thành\s+phố|tp\.?|thị\s+xã|tx\.?|huyện|h\.?|quận|q\.?|thị\s+trấn|tt\.?|phường|p\.?|xã|x\.?|thôn|xóm|ấp|làng|bản)?'
    
    full_pattern = r'\b' + prefix_regex + r'\s*[\.\-–/]?\s*' + core_regex + r'\b'
    
    new_text, count = re.subn(full_pattern, '', text, flags=re.IGNORECASE)
    
    new_text = re.sub(r',\s*,', ',', new_text)
    new_text = re.sub(r'/\s*/', '/', new_text)
    new_text = re.sub(r'-\s*-', '-', new_text)
    new_text = re.sub(r'–\s*–', '–', new_text)
    new_text = re.sub(r'^\s*[\.,\-–/\s]+|[\.,\-–/\s]+$', '', new_text)
    new_text = re.sub(r'\s+', ' ', new_text).strip()
    
    return new_text

def deduplicate_address_parts(parts):
    cleaned = [p.strip() for p in parts if p.strip()]
    for i in range(len(cleaned)):
        for j in range(len(cleaned) - 1, i, -1):
            p_i = cleaned[i]
            p_j = cleaned[j]
            if not p_i or not p_j:
                continue
                
            p_i_norm = normalize_for_comparison(p_i)
            p_j_norm = normalize_for_comparison(p_j)
            
            if p_i_norm == p_j_norm:
                cleaned[i] = ""
                continue
                
            cleaned[i] = remove_target_from_text(p_i, p_j)
            
    return [p for p in cleaned if p.strip()]

def clean_diachi(addr):
    if not addr: return ''
    s = str(addr)
    s = PLUS_CODE_RE.sub('', s)
    s = re.sub(r',\s*,', ',', s)
    s = re.sub(r'^[\s,]+|[\s,]+$', '', s)
    s = re.sub(r'\s+', ' ', s)
    s = s.rstrip('.,;!:')

    parts = [normalize_abbrev(p.strip()) for p in s.split(',')]
    parts = deduplicate_address_parts(parts)
    
    new_parts = []
    for part in parts:
        part = part.strip().rstrip('.,;!:')
        if not part: continue
        part_lower = part.lower()
        matched = None
        for aux in AUX_WORDS:
            if re.match(r'^' + re.escape(aux) + r'(\s|$)', part_lower):
                matched = aux
                break
        if matched:
            aux_cap = matched[0].upper() + matched[1:]
            rest = part[len(matched):].strip()
            rest_cap = ' '.join(title_viet(w) for w in rest.split())
            new_parts.append(aux_cap + (' ' + rest_cap if rest_cap else ''))
        else:
            new_parts.append(' '.join(title_viet(w) for w in part.split()))

    final_parts = []
    for p_val in new_parts:
        p_val_clean = p_val.strip()
        if p_val_clean in ['Hồ Chí Minh', 'Thành Phố Hồ Chí Minh', 'Thành phố Hồ Chí Minh', 'Tp. Hồ Chí Minh', 'Tp.Hồ Chí Minh', 'Tp HCM', 'Tp.HCM']:
            final_parts.append('TP.Hồ Chí Minh')
        else:
            final_parts.append(p_val)

    if not final_parts: return ''
    return ', '.join(final_parts) + '.'

def extract_diaphuong(cleaned_addr):
    if not cleaned_addr: return ''
    s = cleaned_addr.rstrip('.')
    parts = [p.strip() for p in re.split(r'[,\-–.]', s) if p.strip()]
    if not parts: return ''
    
    last_part = parts[-1]
    for aux in AUX_WORDS:
        if re.match(r'^' + re.escape(aux) + r'(\s|$)', last_part, re.IGNORECASE):
            last_part = last_part[len(aux):].strip()
            break

    result = ' '.join(w[0].upper() + w[1:] if w else w for w in last_part.split())
    
    def normalize_hcm(val):
        if val in ['Hồ Chí Minh', 'TP. Hồ Chí Minh', 'TPHCM', 'TP.HCM', 'Thành phố Hồ Chí Minh', 'Thành Phố Hồ Chí Minh']:
            return 'TP.Hồ Chí Minh'
        return val

    if result.lower() in _PROV_LOWER: 
        return normalize_hcm(_PROV_LOWER[result.lower()])

    words = s.split()
    for n in [3, 2, 1]:
        if len(words) >= n:
            candidate = ' '.join(words[-n:])
            if candidate.lower() in _PROV_LOWER:
                return normalize_hcm(_PROV_LOWER[candidate.lower()])
    return normalize_hcm(result)

# ==========================================
# 3. KẾT NỐI API LARK
# ==========================================
def get_tenant_access_token():
    print("🔑 Đang lấy Access Token...")
    url = "https://open.larksuite.com/open-apis/auth/v3/tenant_access_token/internal"
    res = requests.post(url, json={"app_id": APP_ID, "app_secret": APP_SECRET}).json()
    token = res.get('tenant_access_token')
    if not token:
        print("LỖI: Không thể lấy token:", res)
        sys.exit(1)
    return token

def get_field_value(fields, possible_names, default=""):
    for name in possible_names:
        if name in fields:
            val = fields[name]
            if isinstance(val, list) and len(val) > 0:
                return str(val[0].get('text', val[0])).strip()
            return str(val).strip() if val is not None else ""
    return default

def get_file_field_name(fields, possible_names):
    for name in possible_names:
        if name in fields:
            return name
    return FILE_FIELD_NAME

def fetch_and_clean_records(token, table_id, major_name, view_id=None):
    print(f"📥 Bảng {major_name}: Đang lấy dữ liệu thô từ Lark và TỰ ĐỘNG LÀM SẠCH...")
    if view_id:
        print(f"   📋 Lọc theo View ID: {view_id}")
    records = []
    url = f"https://open.larksuite.com/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{table_id}/records"
    headers = {"Authorization": f"Bearer {token}"}
    params = {"page_size": 100}
    if view_id:
        params["view_id"] = view_id
    
    # Định nghĩa các trường tìm kiếm linh hoạt (Self-healing)
    # Ưu tiên tên cột thực tế trên Lark Base trước
    name_fields = ['Họ Và Tên', 'Họ và tên học viên', '*Họ tên HV - Sau lọc', 'Họ tên học viên', 'Họ và tên', 'Họ tên']
    cccd_fields = ['Số CCCD', 'CCCD/ CMT', 'CCCD', 'Số CCCD/CMND', 'CMND/CCCD']
    dob_fields = ['Ngày Sinh', 'Ngày sinh học viên', 'Ngày sinh', 'Ngày sinh HV']
    phone_fields = ['*Số Điện Thoại', 'SĐT học viên', 'SĐT', 'Số điện thoại', 'Điện thoại']
    # Địa chỉ: ưu tiên ghép từ các trường (TT) nếu có
    addr_street_fields = ['(TT)Địa Chỉ Nhà', 'Địa chỉ thường trú', 'Địa chỉ', 'Địa chỉ cụ thể']
    addr_ward_fields = ['(TT)Phường Xã']
    addr_district_fields = ['(TT)Quận Huyện']
    addr_city_fields = ['(TT)Thành Phố']
    loc_fields = ['(TT)Thành Phố', '(HK)Thành Phố', 'Địa Phương', 'Địa phương', 'Nơi sinh']
    email_time_fields = ['Thời gian gửi email đủ ĐK trúng tuyển', 'Thời gian email', 'Thời gian gửi email']
    file_fields = [FILE_FIELD_NAME, 'File trúng tuyển vòng 1', 'File trúng tuyển thí sinh vòng 1', 'File học viên', 'File đính kèm', 'File']

    while True:
        res = requests.get(url, headers=headers, params=params).json()
        if res.get('code') != 0:
            print(f"Lỗi khi lấy records của bảng {major_name}: {res}")
            break
            
        items = res.get('data', {}).get('items', [])
        for item in items:
            fields = item.get('fields', {})
            
            # Trích xuất thông tin sử dụng cơ chế self-healing
            raw_cccd = get_field_value(fields, cccd_fields)
            if not raw_cccd: 
                continue  
            
            raw_hoten = get_field_value(fields, name_fields)
            
            # Xử lý Ngày sinh
            raw_ngaysinh = None
            for f_n in dob_fields:
                if f_n in fields:
                    raw_ngaysinh = fields[f_n]
                    break
            
            raw_sdt = get_field_value(fields, phone_fields)
            
            # Ghép địa chỉ từ các trường (TT) riêng biệt
            raw_street = get_field_value(fields, addr_street_fields)
            raw_ward = get_field_value(fields, addr_ward_fields)
            raw_district = get_field_value(fields, addr_district_fields)
            raw_city = get_field_value(fields, addr_city_fields)
            
            # Lọc sạch địa chỉ cụ thể khỏi các phần trùng lặp với Ward, District, City
            clean_street = clean_street_address(raw_street, raw_ward, raw_district, raw_city)
            
            # Nếu có các trường (TT) riêng, ghép lại thành địa chỉ đầy đủ
            if clean_street or raw_ward or raw_district or raw_city:
                addr_parts = [p for p in [clean_street, raw_ward, raw_district, raw_city] if p and p.strip() and p.strip() != 'N.A']
                raw_diachi = ', '.join(addr_parts)
            else:
                raw_diachi = ''
            
            file_col_name = get_file_field_name(fields, file_fields)
            da_co_file = bool(fields.get(file_col_name))
            
            # Làm sạch dữ liệu
            clean_hoten_val = clean_ten(raw_hoten)
            clean_cccd_val = clean_cccd(raw_cccd)
            clean_ngaysinh_val_str = clean_ngaysinh_val(raw_ngaysinh)
            clean_sdt_val = clean_sdt(raw_sdt)
            clean_diachi_val = clean_diachi(raw_diachi)
            
            # Trích xuất địa phương/nơi sinh
            raw_diaphuong = get_field_value(fields, loc_fields)
            noi_sinh = extract_diaphuong(raw_diaphuong) if raw_diaphuong else extract_diaphuong(clean_diachi_val)
            
            raw_thoigian = get_field_value(fields, email_time_fields)
            
            rec = {
                'record_id': item.get('record_id'),
                'Họ và tên': clean_hoten_val,
                'CCCD': clean_cccd_val,
                'Ngày sinh': clean_ngaysinh_val_str,
                'Nơi sinh': noi_sinh,
                'Địa chỉ': clean_diachi_val,
                'SĐT': clean_sdt_val,
                'Ngành': major_name,
                'Thời gian email': raw_thoigian,
                'Đã có file': da_co_file,
                'file_field_name': file_col_name
            }
            records.append(rec)
                
        if not res.get('data', {}).get('has_more', False):
            break
        params['page_token'] = res.get('data', {}).get('page_token')
        
    print(f"-> Đã lấy và chuẩn hóa thành công {len(records)} hồ sơ từ bảng {major_name}.")
    return records

def create_excel_backup(records, filename_prefix):
    print(f"💾 Đang xuất dữ liệu ra file Excel Backup ({filename_prefix})...")
    groups = {}
    for r in records:
        r_copy = r.copy()
        r_copy.pop('record_id', None)
        r_copy.pop('file_field_name', None)
        
        r_copy['CCCD'] = str(r_copy.get('CCCD', ''))
        r_copy['SĐT'] = str(r_copy.get('SĐT', ''))
        
        tg = r_copy.get('Thời gian email')
        if tg is None:
            tg_key = None
        else:
            tg_key = str(tg).strip()
            if not tg_key or tg_key.lower() in ('nan', 'none'):
                tg_key = None
                
        if tg_key not in groups:
            groups[tg_key] = []
        groups[tg_key].append(r_copy)
        
    for tg_key, group_records in groups.items():
        if tg_key is None:
            file_name = f'Data_Lark_Cleaned_Backup_{filename_prefix}_ChuaCoNgay.xlsx'
        else:
            safe_name_file = re.sub(r'[\\/*?:"<>|]', '-', tg_key)
            file_name = f'Data_Lark_Cleaned_Backup_{filename_prefix}_{safe_name_file}.xlsx'
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Backup"
        
        headers = ['Họ và tên', 'CCCD', 'Ngày sinh', 'Nơi sinh', 'Địa chỉ', 'SĐT', 'Ngành', 'Thời gian email', 'Đã có file']
        ws.append(headers)
        
        # Style định dạng bảng đẹp
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Format Headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        for rec in group_records:
            row = [rec.get(h, '') for h in headers]
            ws.append(row)
            
        # Format dữ liệu và định dạng text cho CCCD/SĐT
        for row in range(2, len(group_records) + 2):
            ws.cell(row=row, column=2).number_format = '@' # Cột CCCD (cột B)
            ws.cell(row=row, column=6).number_format = '@' # Cột SĐT (cột F)
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row, column=col)
                c.border = thin_border
                c.alignment = center_align if col in [2, 3, 6, 7, 8, 9] else left_align
                
        # Tự động chỉnh độ rộng cột
        column_widths = [25, 15, 12, 18, 45, 13, 10, 20, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64+i)].width = width
            
        wb.save(file_name)
        print(f"-> Đã lưu file {file_name} ({len(group_records)} dòng)")

def upload_file_to_lark(token, file_name, file_bytes):
    url = "https://open.larksuite.com/open-apis/drive/v1/medias/upload_all"
    headers = {"Authorization": f"Bearer {token}"}
    files = {'file': (file_name, file_bytes, 'application/pdf')}
    data = {
        'file_name': file_name,
        'parent_type': 'bitable_file',
        'parent_node': APP_TOKEN,
        'size': len(file_bytes)
    }
    
    res = requests.post(url, headers=headers, data=data, files=files).json()
    if res.get('code') != 0:
        print(f" [!] Lỗi upload {file_name}: {res}")
        return None
    return res.get('data', {}).get('file_token')

def update_bitable_record(token, table_id, record_id, file_field_name, file_token):
    url = f"https://open.larksuite.com/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{table_id}/records/{record_id}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        clear_payload = {"fields": {file_field_name: []}}
        requests.put(url, headers=headers, json=clear_payload)
    except Exception:
        pass
    payload = {"fields": {file_field_name: [{"file_token": file_token}]}}
    res = requests.put(url, headers=headers, json=payload).json()
    return res.get('code') == 0

# ==========================================
# 4. HÀM TRỘN THƯ BẰNG PYTHON-DOCX
# ==========================================
def fill_student_info(template_path, student_data, output_path):
    doc = docx.Document(template_path)
    
    for p in doc.paragraphs:
        runs = p.runs
        found = False
        
        # 3 runs check
        for idx in range(len(runs) - 2):
            combined = (runs[idx].text + runs[idx+1].text + runs[idx+2].text).replace(' ', '').lower()
            if combined == "bìnhđịnh" or combined == "tháibình":
                runs[idx].text = student_data['Nơi sinh']
                runs[idx].bold = True
                runs[idx].italic = True
                runs[idx+1].text = ""
                runs[idx+2].text = ""
                found = True
                break
                
        # 2 runs check
        if not found:
            for idx in range(len(runs) - 1):
                combined = (runs[idx].text + runs[idx+1].text).replace(' ', '').lower()
                if combined == "bìnhđịnh" or combined == "tháibình":
                    runs[idx].text = student_data['Nơi sinh']
                    runs[idx].bold = True
                    runs[idx].italic = True
                    runs[idx+1].text = ""
                    found = True
                    break
                    
        # 1 run check
        if not found:
            for idx in range(len(runs)):
                val = runs[idx].text.strip()
                if val == "Bình Định" or val == "Thái Bình":
                    runs[idx].text = student_data['Nơi sinh']
                    runs[idx].bold = True
                    runs[idx].italic = True
                    break
                    
        # Other replacements
        for r in runs:
            if "Nguyễn Phước Nghĩa" in r.text:
                r.text = r.text.replace("Nguyễn Phước Nghĩa", student_data['Họ và tên'])
                r.bold = True
            elif "Bùi Hà Phương" in r.text:
                r.text = r.text.replace("Bùi Hà Phương", student_data['Họ và tên'])
                r.bold = True
            elif "049208001118" in r.text:
                r.text = r.text.replace("049208001118", student_data['CCCD'])
                r.bold = True
                r.italic = True
            elif "034308004046" in r.text:
                r.text = r.text.replace("034308004046", student_data['CCCD'])
                r.bold = True
                r.italic = True
            elif "13/06/2008" in r.text:
                r.text = r.text.replace("13/06/2008", student_data['Ngày sinh'])
                r.bold = True
                r.italic = True
            elif "02/02/2008" in r.text:
                r.text = r.text.replace("02/02/2008", student_data['Ngày sinh'])
                r.bold = True
                r.italic = True
            elif "0797982118" in r.text:
                r.text = r.text.replace("0797982118", student_data['SĐT'])
                r.bold = True
                r.italic = True
            elif "0344637351" in r.text:
                r.text = r.text.replace("0344637351", student_data['SĐT'])
                r.bold = True
                r.italic = True
            elif "28 Đinh Núp, Tổ 5, Phường Kon Tum, Quảng Ngãi." in r.text:
                r.text = " " + student_data['Địa chỉ'].strip()
                r.bold = True
                r.italic = True
            elif "Xóm 3, Thôn Mỹ Bổng, Xã Vạn Xuân, Tỉnh Hưng Yên." in r.text:
                r.text = " " + student_data['Địa chỉ'].strip()
                r.bold = True
                r.italic = True
                
        # Squeeze whitespace for Birthplace paragraph to prevent line wrapping
        if "Nơi sinh:" in p.text and ("Số điện thoại:" in p.text or "SĐT" in p.text):
            for r in runs:
                if r.text.strip() == "" and " " in r.text and len(r.text) > 1:
                    r.text = " "
                    
    doc.save(output_path)

# ==========================================
# 5. CHẠY PIPELINE
# ==========================================
def run_automation():
    token = get_tenant_access_token()
    
    # Lấy dữ liệu từ cả 2 bảng
    records_cntt = fetch_and_clean_records(token, TABLE_CNTT, "CNTT", VIEW_CNTT)
    records_qtkd = fetch_and_clean_records(token, TABLE_QTKD, "QTKD", VIEW_QTKD)
    
    records = records_cntt + records_qtkd
    
    # Deduplicate records by CCCD to keep one unique record per student
    seen_cccd = set()
    unique_records = []
    for r in records:
        cccd = r.get('CCCD')
        if cccd not in seen_cccd:
            seen_cccd.add(cccd)
            unique_records.append(r)
        else:
            print(f" -> Phát hiện bản ghi trùng lặp CCCD ({cccd}): Bỏ qua bản ghi của {r['Họ và tên']}")
    records = unique_records
    
    if not records:
        print("Không có dữ liệu hợp lệ để xử lý.")
        return
        
    if records_cntt:
        create_excel_backup(records_cntt, "CNTT")
    if records_qtkd:
        create_excel_backup(records_qtkd, "QTKD")
    
    # Chuẩn bị thư mục tạm
    temp_dir = Path("_temp_all_in_one")
    if temp_dir.exists(): 
        shutil.rmtree(temp_dir, ignore_errors=True)
    temp_dir.mkdir()
    
    # Lọc học sinh cần xử lý (SKIP_EXISTING = True cho auto_update_new_app)
    students_to_process = []
    for r in records:
        if SKIP_EXISTING and r['Đã có file']:
            continue
        students_to_process.append(r)
        
    if not students_to_process:
        print("🎉 Tất cả học sinh đều đã có File học viên trên Lark. Không cần chạy thêm.")
        processed_students = []
        for r in records:
            processed_students.append({
                "name": r['Họ và tên'],
                "cccd": r['CCCD'],
                "dob": r['Ngày sinh'],
                "phone": r['SĐT'],
                "major": r['Ngành'],
                "status": "done"
            })
        try:
            with open('processed_students.json', 'w', encoding='utf-8') as f:
                json.dump(processed_students, f, ensure_ascii=False, indent=4)
        except Exception:
            pass
        return
        
    print(f"\n⚙️ Bắt đầu trộn thư (Mail Merge) cho {len(students_to_process)} học sinh...")
    docx_list = []
    
    for r in students_to_process:
        if r['Ngành'] == "CNTT":
            template = "template_CNTT.docx"
        else:
            template = "template_QTKD.docx"
            
        file_name = f"{r['Họ và tên']}_{r['CCCD']}"
        docx_path = temp_dir / f"{file_name}.docx"
        
        fill_student_info(template, r, docx_path)
        
        docx_list.append({
            'record_id': r['record_id'],
            'file_name_base': file_name,
            'docx_path': str(docx_path),
            'nganh': r['Ngành'],
            'file_field_name': r['file_field_name']
        })
        
    print("-> Xong phần trộn thư.")
    
    print("\n📄 Bắt đầu chuyển đổi hàng loạt sang PDF (Mất vài phút)...")
    if sys.platform == "win32":
        convert(str(temp_dir))
    else:
        import subprocess
        docx_files = list(temp_dir.glob("*.docx"))
        total_conv = len(docx_files)
        print(f"Đang dùng LibreOffice để chuyển đổi DOCX -> PDF trên Linux ({total_conv} file)...")
        if total_conv > 0:
            batch_size = 50
            for i in range(0, total_conv, batch_size):
                batch = docx_files[i:i+batch_size]
                print(f"-> Đang convert PDF nhóm {i//batch_size + 1} ({len(batch)} file)...", flush=True)
                cmd = [
                    "libreoffice", "--headless", "--convert-to", "pdf", 
                    "--outdir", str(temp_dir)
                ] + [str(f) for f in batch]
                subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            
            # Kháng lỗi nâng cao (Self-healing)
            missing_files = []
            for f in docx_files:
                pdf_path = temp_dir / f"{f.stem}.pdf"
                if not pdf_path.exists():
                    missing_files.append(f)
            
            if missing_files:
                print(f"⚠ Phát hiện {len(missing_files)} file PDF bị thiếu. Tiến hành convert bổ sung...", flush=True)
                for f in missing_files:
                    subprocess.run([
                        "libreoffice", "--headless", "--convert-to", "pdf", 
                        "--outdir", str(temp_dir), str(f)
                    ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    
            print(f"-> Đã convert thành công toàn bộ {total_conv}/{total_conv} file sang PDF!")
    
    print("\n📦 Đang nén các file PDF vào file ZIP theo ngành...")
    zip_cntt = zipfile.ZipFile("PDF_TrungTuyen_CNTT.zip", 'w', zipfile.ZIP_DEFLATED)
    zip_qtkd = zipfile.ZipFile("PDF_TrungTuyen_QTKD.zip", 'w', zipfile.ZIP_DEFLATED)
    
    for item in docx_list:
        pdf_path = temp_dir / f"{item['file_name_base']}.pdf"
        if pdf_path.exists():
            if item['nganh'] == "QTKD":
                zip_qtkd.write(pdf_path, arcname=pdf_path.name)
            else:
                zip_cntt.write(pdf_path, arcname=pdf_path.name)
                
    zip_cntt.close()
    zip_qtkd.close()
    print("-> Đã tạo 2 file ZIP thành công (CNTT và QTKD) để bạn dự phòng!")
 
    print("\n☁️ Bắt đầu Upload lên Lark Bitable...")
    success_count = 0
    total = len(docx_list)
    
    for idx, item in enumerate(docx_list, 1):
        pdf_path = temp_dir / f"{item['file_name_base']}.pdf"
        print(f"[{idx}/{total}] Đang tải lên {pdf_path.name} ...", end="", flush=True)
        
        item['status'] = 'failure'
        
        if not pdf_path.exists():
            print(" LỖI: Không tìm thấy file PDF!")
            continue
            
        with open(pdf_path, 'rb') as f:
            file_bytes = f.read()
            
        try:
            token_file = upload_file_to_lark(token, pdf_path.name, file_bytes)
            if token_file:
                target_table = TABLE_CNTT if item['nganh'] == "CNTT" else TABLE_QTKD
                if update_bitable_record(token, target_table, item['record_id'], item['file_field_name'], token_file):
                    print(" Thành công! ✓")
                    success_count += 1
                    item['status'] = 'success'
                else:
                    print(" LỖI Cập nhật!")
            else:
                print(" LỖI Upload!")
        except Exception as e:
            print(f" LỖI Kết nối: {e}")
            continue
            
    # Xuất kết quả
    processed_students = []
    for item in docx_list:
        student_info = next((rec for rec in records if rec['record_id'] == item['record_id']), None)
        if student_info:
            processed_students.append({
                "name": student_info['Họ và tên'],
                "cccd": student_info['CCCD'],
                "dob": student_info['Ngày sinh'],
                "phone": student_info['SĐT'],
                "major": student_info['Ngành'],
                "status": "done" if item['status'] == 'success' else "error"
            })
            
    # Thêm những học sinh cũ
    for r in records:
        if SKIP_EXISTING and r['Đã có file']:
            processed_students.append({
                "name": r['Họ và tên'],
                "cccd": r['CCCD'],
                "dob": r['Ngày sinh'],
                "phone": r['SĐT'],
                "major": r['Ngành'],
                "status": "done"
            })
            
    try:
        with open('processed_students.json', 'w', encoding='utf-8') as f:
            json.dump(processed_students, f, ensure_ascii=False, indent=4)
        print("-> Đã lưu kết quả xử lý thực tế ra file processed_students.json!")
    except Exception as e:
        print(f"Lỗi ghi file processed_students.json: {e}")
                
    print(f"\n🧹 Đang dọn dẹp thư mục tạm...")
    shutil.rmtree(temp_dir, ignore_errors=True)
    
    print(f"✅ HOÀN TẤT ALL-IN-ONE PIPELINE! Đã đẩy thành công {success_count}/{total} file.")

if __name__ == "__main__":
    run_automation()
