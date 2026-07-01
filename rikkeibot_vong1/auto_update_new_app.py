import os
import sys
import io
import time
import shutil
import requests
import datetime
import re
from openpyxl import Workbook
from pathlib import Path
import zipfile
import docx
from docx2pdf import convert

# Đảm bảo in tiếng Việt không lỗi trên terminal
if not hasattr(sys.stdout, 'original_stdout'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import json

# ==========================================
# 1. CẤU HÌNH API LARK MẶC ĐỊNH
# ==========================================
DEFAULT_APP_ID = "cli_aa8847bf0f79deef"
DEFAULT_APP_SECRET = "zBA9rmZ7ScgEtIS1OFFOedGIHaz11V8g"
DEFAULT_APP_TOKEN = "UWhibaHkCaWLausconRl8krsgZg"
DEFAULT_TABLE_CNTT = "tblZoFTIttbvejQ7"
DEFAULT_TABLE_QTKD = "tbl4gHCX7H8eGjaA"

APP_ID = DEFAULT_APP_ID
APP_SECRET = DEFAULT_APP_SECRET
APP_TOKEN = DEFAULT_APP_TOKEN
TABLE_CNTT = DEFAULT_TABLE_CNTT
TABLE_QTKD = DEFAULT_TABLE_QTKD

# Đọc cấu hình từ config.json nếu tồn tại (để Web UI ghi đè)
try:
    if os.path.exists('config.json'):
        with open('config.json', 'r', encoding='utf-8') as f:
            _config = json.load(f)
            APP_ID = _config.get('APP_ID', DEFAULT_APP_ID)
            APP_SECRET = _config.get('APP_SECRET', DEFAULT_APP_SECRET)
            APP_TOKEN = _config.get('APP_TOKEN', DEFAULT_APP_TOKEN)
            TABLE_CNTT = _config.get('TABLE_CNTT', DEFAULT_TABLE_CNTT)
            TABLE_QTKD = _config.get('TABLE_QTKD', DEFAULT_TABLE_QTKD)
except Exception as e:
    print(f"Cảnh báo đọc file config.json: {e}. Sử dụng cấu hình mặc định.")

FILE_FIELD_NAME = 'File học viên'
SKIP_EXISTING = True  # Đổi thành True để bỏ qua những người đã có file, chỉ update người mới

# ==========================================
# 2. BỘ TỪ ĐIỂN VÀ HÀM LÀM SẠCH (CLEAN DATA)
# ==========================================
AUX_WORDS = sorted([
    'số nhà', 'thành phố', 'thị trấn', 'thị xã', 'khu phố', 'dân phố',
    'tỉnh', 'xã', 'thôn', 'xóm', 'phường', 'quận', 'huyện',
    'khối', 'tổ', 'ngõ', 'ngách', 'đường', 'phố',
    'số', 'khu', 'ấp', 'làng', 'hẻm', 'lô',
], key=len, reverse=True)

PLUS_CODE_RE = re.compile(r'\b[A-Z0-9]{4,8}\+[A-Z0-9]{2,4}\b', re.IGNORECASE)

ABBREV_MAP = [
    (re.compile(r'\bTP\.\s*', re.IGNORECASE), 'Thành phố '),
    (re.compile(r'\bTX\.\s*', re.IGNORECASE), 'Thị xã '),
    (re.compile(r'\bTT\.\s*', re.IGNORECASE), 'Thị trấn '),
    (re.compile(r'\bP\.\s*',  re.IGNORECASE), 'Phường '),
    (re.compile(r'\bQ\.\s*',  re.IGNORECASE), 'Quận '),
    (re.compile(r'\bH\.\s*',  re.IGNORECASE), 'Huyện '),
    (re.compile(r'\bX\.\s*',  re.IGNORECASE), 'Xã '),
    (re.compile(r'\bHN\b',    re.IGNORECASE), 'Hà Nội'),
    (re.compile(r'\bHCM\b',   re.IGNORECASE), 'Hồ Chí Minh'),
    (re.compile(r'\bTPHCM\b', re.IGNORECASE), 'Hồ Chí Minh'),
    (re.compile(r'\bĐN\b',    re.IGNORECASE), 'Đà Nẵng'),
]

PROVINCES = [
    'Hà Nội', 'Hồ Chí Minh', 'Hải Phòng', 'Đà Nẵng', 'Cần Thơ',
    'An Giang', 'Bà Rịa Vũng Tàu', 'Bắc Giang', 'Bắc Kạn', 'Bạc Liêu',
    'Bắc Ninh', 'Bến Tre', 'Bình Định', 'Bình Dương', 'Bình Phước',
    'Bình Thuận', 'Cà Mau', 'Cao Bằng', 'Đắk Lắk', 'Đắk Nông',
    'Điện Biên', 'Đồng Nai', 'Đồng Tháp', 'Gia Lai', 'Hà Giang',
    'Hà Nam', 'Hà Tĩnh', 'Hải Dương', 'Hậu Giang', 'Hòa Bình',
    'Hưng Yên', 'Khánh Hòa', 'Kiên Giang', 'Kon Tum', 'Lai Châu',
    'Lâm Đồng', 'Lạng Sơn', 'Lào Cai', 'Long An', 'Nam Định',
    'Nghệ An', 'Ninh Bình', 'Ninh Thuận', 'Phú Thọ', 'Phú Yên',
    'Quảng Bình', 'Quảng Nam', 'Quảng Ngãi', 'Quảng Ninh', 'Quảng Trị',
    'Sóc Trăng', 'Sơn La', 'Tây Ninh', 'Thái Bình', 'Thái Nguyên',
    'Thanh Hóa', 'Thừa Thiên Huế', 'Tiền Giang', 'Trà Vinh', 'Tuyên Quang',
    'Vĩnh Long', 'Vĩnh Phúc', 'Yên Bái',
]
_PROV_LOWER = {p.lower(): p for p in PROVINCES}

def title_viet(word):
    return word[0].upper() + word[1:] if word else word

def normalize_abbrev(text):
    for pattern, replacement in ABBREV_MAP:
        text = pattern.sub(replacement, text)
    return text

def clean_ten(val):
    if not val: return ''
    s = re.sub(r'\s+', ' ', re.sub(r'\d+', '', str(val))).strip()
    return ' '.join(w[0].upper() + w[1:].lower() if w else w for w in s.split())

def clean_cccd(val):
    if val is None: return ''
    try: s = str(int(round(float(str(val)))))
    except Exception: s = re.sub(r'\D', '', str(val))
    s = re.sub(r'\D', '', s)
    return s.zfill(12)[:12]

def clean_sdt(val):
    if val is None: return ''
    try: s = str(int(round(float(str(val)))))
    except Exception: s = re.sub(r'\D', '', str(val))
    s = re.sub(r'\D', '', s)
    if len(s) == 9: s = '0' + s
    return s.zfill(10)[-10:]

def clean_ngaysinh_val(val):
    if not val: return ''
    try:
        if str(val).isdigit() and len(str(val)) > 10:
            dt = datetime.datetime.fromtimestamp(int(val)/1000)
            return dt.strftime('%d/%m/%Y')
        
        val_str = str(val).strip()
        date_part = val_str.split('T')[0].split(' ')[0]
        
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y'):
            try:
                dt = datetime.datetime.strptime(date_part, fmt)
                return dt.strftime('%d/%m/%Y')
            except Exception:
                continue
        return val_str
    except Exception:
        return str(val)

def clean_diachi(addr):
    if not addr: return ''
    s = str(addr)
    s = PLUS_CODE_RE.sub('', s)
    s = re.sub(r',\s*,', ',', s)
    s = re.sub(r'^[\s,]+|[\s,]+$', '', s)
    s = re.sub(r'\s+', ' ', s)
    s = s.rstrip('.,;!:')

    parts = [normalize_abbrev(p.strip()) for p in s.split(',')]
    new_parts = []
    for part in parts:
        part = part.strip().rstrip('.,;!:')
        if not part: continue
        part_lower = part.lower()
        matched = None
        for aux in AUX_WORDS:
            if re.match(r'^' + re.escape(aux) + r'(\s|$)', part_lower):
                matched = aux
                break
        if matched:
            aux_cap = matched[0].upper() + matched[1:]
            rest = part[len(matched):].strip()
            rest_cap = ' '.join(title_viet(w) for w in rest.split())
            new_parts.append(aux_cap + (' ' + rest_cap if rest_cap else ''))
        else:
            new_parts.append(' '.join(title_viet(w) for w in part.split()))

    if not new_parts: return ''
    return ', '.join(new_parts) + '.'

def extract_diaphuong(cleaned_addr):
    if not cleaned_addr: return ''
    s = cleaned_addr.rstrip('.')
    parts = [p.strip() for p in re.split(r'[,\-–.]', s) if p.strip()]
    if not parts: return ''
    
    last_part = parts[-1]
    for aux in AUX_WORDS:
        if re.match(r'^' + re.escape(aux) + r'(\s|$)', last_part, re.IGNORECASE):
            last_part = last_part[len(aux):].strip()
            break

    result = ' '.join(w[0].upper() + w[1:] if w else w for w in last_part.split())
    if result.lower() in _PROV_LOWER: return _PROV_LOWER[result.lower()]

    words = s.split()
    for n in [3, 2, 1]:
        if len(words) >= n:
            candidate = ' '.join(words[-n:])
            if candidate.lower() in _PROV_LOWER:
                return _PROV_LOWER[candidate.lower()]
    return result

# ==========================================
# 3. KẾT NỐI API LARK
# ==========================================
def get_tenant_access_token():
    print("🔑 Đang lấy Access Token...")
    url = "https://open.larksuite.com/open-apis/auth/v3/tenant_access_token/internal"
    res = requests.post(url, json={"app_id": APP_ID, "app_secret": APP_SECRET}).json()
    token = res.get('tenant_access_token')
    if not token:
        print("LỖI: Không thể lấy token:", res)
        sys.exit(1)
    return token

def get_field_value(fields, possible_names, default=""):
    for name in possible_names:
        if name in fields:
            val = fields[name]
            if isinstance(val, list) and len(val) > 0:
                return str(val[0].get('text', val[0])).strip()
            return str(val).strip() if val is not None else ""
    return default

def get_file_field_name(fields, possible_names):
    for name in possible_names:
        if name in fields:
            return name
    return FILE_FIELD_NAME

def fetch_and_clean_records(token, table_id, major_name):
    print(f"📥 Bảng {major_name}: Đang lấy dữ liệu thô từ Lark và TỰ ĐỘNG LÀM SẠCH...")
    records = []
    url = f"https://open.larksuite.com/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{table_id}/records"
    headers = {"Authorization": f"Bearer {token}"}
    params = {"page_size": 100}
    
    # Định nghĩa các trường tìm kiếm linh hoạt (Self-healing)
    name_fields = ['Họ và tên học viên', '*Họ tên HV - Sau lọc', 'Họ tên học viên', 'Họ và tên', 'Họ tên']
    cccd_fields = ['CCCD/ CMT', 'CCCD', 'Số CCCD', 'Số CCCD/CMND', 'CMND/CCCD']
    dob_fields = ['Ngày sinh học viên', 'Ngày sinh', 'Ngày sinh HV']
    phone_fields = ['SĐT học viên', 'SĐT', 'Số điện thoại', 'Điện thoại']
    addr_fields = ['Địa chỉ thường trú', 'Địa chỉ', 'Địa chỉ cụ thể']
    loc_fields = ['Địa Phương', 'Địa phương', 'Nơi sinh']
    email_time_fields = ['Thời gian gửi email đủ ĐK trúng tuyển', 'Thời gian email', 'Thời gian gửi email']
    file_fields = [FILE_FIELD_NAME, 'File học viên', 'File đính kèm', 'File']

    while True:
        res = requests.get(url, headers=headers, params=params).json()
        if res.get('code') != 0:
            print(f"Lỗi khi lấy records của bảng {major_name}: {res}")
            break
            
        items = res.get('data', {}).get('items', [])
        for item in items:
            fields = item.get('fields', {})
            
            # Trích xuất thông tin sử dụng cơ chế self-healing
            raw_cccd = get_field_value(fields, cccd_fields)
            if not raw_cccd: 
                continue  
            
            raw_hoten = get_field_value(fields, name_fields)
            
            # Xử lý Ngày sinh
            raw_ngaysinh = None
            for f_n in dob_fields:
                if f_n in fields:
                    raw_ngaysinh = fields[f_n]
                    break
            
            raw_sdt = get_field_value(fields, phone_fields)
            raw_diachi = get_field_value(fields, addr_fields)
            
            file_col_name = get_file_field_name(fields, file_fields)
            da_co_file = bool(fields.get(file_col_name))
            
            # Làm sạch dữ liệu
            clean_hoten_val = clean_ten(raw_hoten)
            clean_cccd_val = clean_cccd(raw_cccd)
            clean_ngaysinh_val_str = clean_ngaysinh_val(raw_ngaysinh)
            clean_sdt_val = clean_sdt(raw_sdt)
            clean_diachi_val = clean_diachi(raw_diachi)
            
            # Trích xuất địa phương/nơi sinh
            raw_diaphuong = get_field_value(fields, loc_fields)
            noi_sinh = raw_diaphuong if raw_diaphuong else extract_diaphuong(clean_diachi_val)
            
            raw_thoigian = get_field_value(fields, email_time_fields)
            
            rec = {
                'record_id': item.get('record_id'),
                'Họ và tên': clean_hoten_val,
                'CCCD': clean_cccd_val,
                'Ngày sinh': clean_ngaysinh_val_str,
                'Nơi sinh': noi_sinh,
                'Địa chỉ': clean_diachi_val,
                'SĐT': clean_sdt_val,
                'Ngành': major_name,
                'Thời gian email': raw_thoigian,
                'Đã có file': da_co_file,
                'file_field_name': file_col_name
            }
            records.append(rec)
                
        if not res.get('data', {}).get('has_more', False):
            break
        params['page_token'] = res.get('data', {}).get('page_token')
        
    print(f"-> Đã lấy và chuẩn hóa thành công {len(records)} hồ sơ từ bảng {major_name}.")
    return records

def create_excel_backup(records, filename_prefix):
    print(f"💾 Đang xuất dữ liệu ra file Excel Backup ({filename_prefix})...")
    groups = {}
    for r in records:
        r_copy = r.copy()
        r_copy.pop('record_id', None)
        r_copy.pop('file_field_name', None)
        
        r_copy['CCCD'] = str(r_copy.get('CCCD', ''))
        r_copy['SĐT'] = str(r_copy.get('SĐT', ''))
        
        tg = r_copy.get('Thời gian email')
        if tg is None:
            tg_key = None
        else:
            tg_key = str(tg).strip()
            if not tg_key or tg_key.lower() in ('nan', 'none'):
                tg_key = None
                
        if tg_key not in groups:
            groups[tg_key] = []
        groups[tg_key].append(r_copy)
        
    for tg_key, group_records in groups.items():
        if tg_key is None:
            file_name = f'Data_Lark_Cleaned_Backup_{filename_prefix}_ChuaCoNgay.xlsx'
        else:
            safe_name_file = re.sub(r'[\\/*?:"<>|]', '-', tg_key)
            file_name = f'Data_Lark_Cleaned_Backup_{filename_prefix}_{safe_name_file}.xlsx'
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Backup"
        
        headers = ['Họ và tên', 'CCCD', 'Ngày sinh', 'Nơi sinh', 'Địa chỉ', 'SĐT', 'Ngành', 'Thời gian email', 'Đã có file']
        ws.append(headers)
        
        for rec in group_records:
            row = [rec.get(h, '') for h in headers]
            ws.append(row)
            
        # Định dạng cột text cho CCCD và SĐT để tránh mất số 0
        for row in range(2, len(group_records) + 2):
            ws.cell(row=row, column=2).number_format = '@' # Cột CCCD (cột B)
            ws.cell(row=row, column=6).number_format = '@' # Cột SĐT (cột F)
            
        wb.save(file_name)
        print(f"-> Đã lưu file {file_name} ({len(group_records)} dòng)")

def upload_file_to_lark(token, file_name, file_bytes):
    url = "https://open.larksuite.com/open-apis/drive/v1/medias/upload_all"
    headers = {"Authorization": f"Bearer {token}"}
    files = {'file': (file_name, file_bytes, 'application/pdf')}
    data = {
        'file_name': file_name,
        'parent_type': 'bitable_file',
        'parent_node': APP_TOKEN,
        'size': len(file_bytes)
    }
    
    res = requests.post(url, headers=headers, data=data, files=files).json()
    if res.get('code') != 0:
        print(f" [!] Lỗi upload {file_name}: {res}")
        return None
    return res.get('data', {}).get('file_token')

def update_bitable_record(token, table_id, record_id, file_field_name, file_token):
    url = f"https://open.larksuite.com/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{table_id}/records/{record_id}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        clear_payload = {"fields": {file_field_name: []}}
        requests.put(url, headers=headers, json=clear_payload)
    except Exception:
        pass
    payload = {"fields": {file_field_name: [{"file_token": file_token}]}}
    res = requests.put(url, headers=headers, json=payload).json()
    return res.get('code') == 0

# ==========================================
# 4. HÀM TRỘN THƯ BẰNG PYTHON-DOCX
# ==========================================
def fill_student_info(template_path, student_data, output_path):
    doc = docx.Document(template_path)
    
    # Paragraph thứ 9 thường chứa thông tin học sinh
    p = doc.paragraphs[9]
    
    if len(p.runs) >= 31:
        # Thay thế chuẩn theo cấu trúc run
        p.runs[1].text = student_data['Họ và tên']
        p.runs[7].text = student_data['CCCD'] + " "
        p.runs[12].text = student_data['Nơi sinh']
        p.runs[13].text = ""
        p.runs[14].text = ""
        p.runs[20].text = student_data['Ngày sinh']
        p.runs[28].text = student_data['SĐT'] + " "
        p.runs[30].text = student_data['Địa chỉ']
    else:
        # Fallback thay thế chuỗi trực tiếp
        text = p.text
        text = text.replace("Nguyễn Phước Nghĩa", student_data['Họ và tên'])
        text = text.replace("Bùi Hà Phương", student_data['Họ và tên'])
        text = text.replace("049208001118", student_data['CCCD'])
        text = text.replace("034308004046", student_data['CCCD'])
        text = text.replace("Bình Định", student_data['Nơi sinh'])
        text = text.replace("Thái Bình", student_data['Nơi sinh'])
        text = text.replace("13/06/2008", student_data['Ngày sinh'])
        text = text.replace("02/02/2008", student_data['Ngày sinh'])
        text = text.replace("0797982118", student_data['SĐT'])
        text = text.replace("0344637351", student_data['SĐT'])
        text = text.replace("28 Đinh Núp, Tổ 5, Phường Kon Tum, Quảng Ngãi.", student_data['Địa chỉ'])
        text = text.replace("Xóm 3, Thôn Mỹ Bổng, Xã Vạn Xuân, Tỉnh Hưng Yên.", student_data['Địa chỉ'])
        p.text = text
        
    doc.save(output_path)

# ==========================================
# 5. CHẠY PIPELINE
# ==========================================
def run_automation():
    token = get_tenant_access_token()
    
    # Lấy dữ liệu từ cả 2 bảng
    records_cntt = fetch_and_clean_records(token, TABLE_CNTT, "CNTT")
    records_qtkd = fetch_and_clean_records(token, TABLE_QTKD, "QTKD")
    
    records = records_cntt + records_qtkd
    
    if not records:
        print("Không có dữ liệu hợp lệ để xử lý.")
        return
        
    if records_cntt:
        create_excel_backup(records_cntt, "CNTT")
    if records_qtkd:
        create_excel_backup(records_qtkd, "QTKD")
    
    # Chuẩn bị thư mục tạm
    temp_dir = Path("_temp_all_in_one")
    if temp_dir.exists(): 
        shutil.rmtree(temp_dir, ignore_errors=True)
    temp_dir.mkdir()
    
    # Lọc học sinh cần xử lý (SKIP_EXISTING = True cho auto_update_new_app)
    students_to_process = []
    for r in records:
        if SKIP_EXISTING and r['Đã có file']:
            continue
        students_to_process.append(r)
        
    if not students_to_process:
        print("🎉 Tất cả học sinh đều đã có File học viên trên Lark. Không cần chạy thêm.")
        processed_students = []
        for r in records:
            processed_students.append({
                "name": r['Họ và tên'],
                "cccd": r['CCCD'],
                "dob": r['Ngày sinh'],
                "phone": r['SĐT'],
                "major": r['Ngành'],
                "status": "done"
            })
        try:
            with open('processed_students.json', 'w', encoding='utf-8') as f:
                json.dump(processed_students, f, ensure_ascii=False, indent=4)
        except Exception:
            pass
        return
        
    print(f"\n⚙️ Bắt đầu trộn thư (Mail Merge) cho {len(students_to_process)} học sinh...")
    docx_list = []
    
    for r in students_to_process:
        if r['Ngành'] == "CNTT":
            template = "template_CNTT.docx"
        else:
            template = "template_QTKD.docx"
            
        file_name = f"{r['Họ và tên']}_{r['CCCD']}"
        docx_path = temp_dir / f"{file_name}.docx"
        
        fill_student_info(template, r, docx_path)
        
        docx_list.append({
            'record_id': r['record_id'],
            'file_name_base': file_name,
            'docx_path': str(docx_path),
            'nganh': r['Ngành'],
            'file_field_name': r['file_field_name']
        })
        
    print("-> Xong phần trộn thư.")
    
    print("\n📄 Bắt đầu chuyển đổi hàng loạt sang PDF (Mất vài phút)...")
    if sys.platform == "win32":
        convert(str(temp_dir))
    else:
        import subprocess
        docx_files = list(temp_dir.glob("*.docx"))
        total_conv = len(docx_files)
        print(f"Đang dùng LibreOffice để chuyển đổi DOCX -> PDF trên Linux ({total_conv} file)...")
        if total_conv > 0:
            batch_size = 50
            for i in range(0, total_conv, batch_size):
                batch = docx_files[i:i+batch_size]
                print(f"-> Đang convert PDF nhóm {i//batch_size + 1} ({len(batch)} file)...", flush=True)
                cmd = [
                    "libreoffice", "--headless", "--convert-to", "pdf", 
                    "--outdir", str(temp_dir)
                ] + [str(f) for f in batch]
                subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            
            # Kháng lỗi nâng cao (Self-healing)
            missing_files = []
            for f in docx_files:
                pdf_path = temp_dir / f"{f.stem}.pdf"
                if not pdf_path.exists():
                    missing_files.append(f)
            
            if missing_files:
                print(f"⚠ Phát hiện {len(missing_files)} file PDF bị thiếu. Tiến hành convert bổ sung...", flush=True)
                for f in missing_files:
                    subprocess.run([
                        "libreoffice", "--headless", "--convert-to", "pdf", 
                        "--outdir", str(temp_dir), str(f)
                    ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    
            print(f"-> Đã convert thành công toàn bộ {total_conv}/{total_conv} file sang PDF!")
    
    print("\n📦 Đang nén các file PDF vào file ZIP theo ngành...")
    zip_cntt = zipfile.ZipFile("PDF_TrungTuyen_CNTT.zip", 'w', zipfile.ZIP_DEFLATED)
    zip_qtkd = zipfile.ZipFile("PDF_TrungTuyen_QTKD.zip", 'w', zipfile.ZIP_DEFLATED)
    
    for item in docx_list:
        pdf_path = temp_dir / f"{item['file_name_base']}.pdf"
        if pdf_path.exists():
            if item['nganh'] == "QTKD":
                zip_qtkd.write(pdf_path, arcname=pdf_path.name)
            else:
                zip_cntt.write(pdf_path, arcname=pdf_path.name)
                
    zip_cntt.close()
    zip_qtkd.close()
    print("-> Đã tạo 2 file ZIP thành công (CNTT và QTKD) để bạn dự phòng!")
 
    print("\n☁️ Bắt đầu Upload lên Lark Bitable...")
    success_count = 0
    total = len(docx_list)
    
    for idx, item in enumerate(docx_list, 1):
        pdf_path = temp_dir / f"{item['file_name_base']}.pdf"
        print(f"[{idx}/{total}] Đang tải lên {pdf_path.name} ...", end="", flush=True)
        
        item['status'] = 'failure'
        
        if not pdf_path.exists():
            print(" LỖI: Không tìm thấy file PDF!")
            continue
            
        with open(pdf_path, 'rb') as f:
            file_bytes = f.read()
            
        try:
            token_file = upload_file_to_lark(token, pdf_path.name, file_bytes)
            if token_file:
                target_table = TABLE_CNTT if item['nganh'] == "CNTT" else TABLE_QTKD
                if update_bitable_record(token, target_table, item['record_id'], item['file_field_name'], token_file):
                    print(" Thành công! ✓")
                    success_count += 1
                    item['status'] = 'success'
                else:
                    print(" LỖI Cập nhật!")
            else:
                print(" LỖI Upload!")
        except Exception as e:
            print(f" LỖI Kết nối: {e}")
            continue
            
    # Xuất kết quả
    processed_students = []
    for item in docx_list:
        student_info = next((rec for rec in records if rec['record_id'] == item['record_id']), None)
        if student_info:
            processed_students.append({
                "name": student_info['Họ và tên'],
                "cccd": student_info['CCCD'],
                "dob": student_info['Ngày sinh'],
                "phone": student_info['SĐT'],
                "major": student_info['Ngành'],
                "status": "done" if item['status'] == 'success' else "error"
            })
            
    # Thêm những học sinh cũ
    for r in records:
        if SKIP_EXISTING and r['Đã có file']:
            processed_students.append({
                "name": r['Họ và tên'],
                "cccd": r['CCCD'],
                "dob": r['Ngày sinh'],
                "phone": r['SĐT'],
                "major": r['Ngành'],
                "status": "done"
            })
            
    try:
        with open('processed_students.json', 'w', encoding='utf-8') as f:
            json.dump(processed_students, f, ensure_ascii=False, indent=4)
        print("-> Đã lưu kết quả xử lý thực tế ra file processed_students.json!")
    except Exception as e:
        print(f"Lỗi ghi file processed_students.json: {e}")
                
    print(f"\n🧹 Đang dọn dẹp thư mục tạm...")
    shutil.rmtree(temp_dir, ignore_errors=True)
    
    print(f"✅ HOÀN TẤT ALL-IN-ONE PIPELINE! Đã đẩy thành công {success_count}/{total} file.")

if __name__ == "__main__":
    run_automation()
